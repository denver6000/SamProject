import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import re
import queue
import threading
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class StudentInfoSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Student Information System")
        self.root.geometry("800x600")
        self.root.minsize(600, 400)  # Minimum window size
        self.root.configure(bg="white")  # White background
        self.data_file = 'student_data.json'
        self.trash_file = 'deleted_students.json'
        self.options_file = 'school_course_options.json'
        
        # Initialize school and course options
        self.school_options = [
            'PHINMA Araullo University San Jose City ',
            'Core Gateway College, Inc.',
            'San Jose Christian Colleges',
            'Golden Success University',
            'STI College San Jose',
            'Central Luzon State University',
            'Others'
        ]
        
        self.course_mapping = {
            'PHINMA Araullo University San Jose City ': [
                'Accountancy', 'Business Administration', 'Hospitality Management', 'Criminology',
                'Information Technology', 'Computer Science', 'Education (Elementary & Secondary)',
                'Nursing', 'Civil Engineering'
            ],
            'Core Gateway College, Inc.': [
                'Computer Science', 'Business Administration (Management, Accounting)', 'Political Science',
                'Education (Elementary & Secondary)', "Master's: Education, Public Administration"
            ],
            'San Jose Christian Colleges': [
                'Business Administration', 'Education', 'Criminology', 'Information Technology',
                'Hospitality / Tourism'
            ],
            'Golden Success University': [
                'Business Administration', 'Education', 'Information Technology', 'Criminology'
            ],
            'STI College San Jose': [
                'Information Technology', 'Computer Science', 'Hospitality Management',
                'Tourism Management', 'Business Administration'
            ],
            'Central Luzon State University': [
                'BS Agriculture',
                'BS Agribusiness',
                'BS Agricultural and Biosystems Engineering',
                'BS Animal Science / Animal Husbandry',
                'BS Fisheries / Aquaculture',
                'BS Biology',
                'BS Chemistry',
                'BS Environmental Science',
                'BS Civil Engineering',
                'BS Information Technology',
                'BS Accountancy',
                'BS Accounting Technology',
                'BS Business Administration',
                'BS Entrepreneurship',
                'Bachelor of Elementary Education (BEEd)',
                'Bachelor of Secondary Education (BSEd) – English',
                'Bachelor of Secondary Education (BSEd) – Math',
                'Bachelor of Secondary Education (BSEd) – Science',
                'Bachelor of Secondary Education (BSEd) – Filipino',
                'Bachelor of Secondary Education (BSEd) – Social Studies',
                'Bachelor of Secondary Education (BSEd) – MAPEH',
                'Bachelor of Secondary Education (BSEd) – TLE',
                'Bachelor of Early Childhood Education',
                'Bachelor of Physical Education',
                'BA Psychology',
                'BA Social Sciences',
                'BA Filipino',
                'BA Language / Literature',
                'BS Hospitality Management',
                'BS Food Technology',
                'BS Fashion and Textile Technology',
                'BS Development Communication',
                'Doctor of Veterinary Medicine (DVM)'
            ],
            'Others': []  # Will be populated dynamically
        }
        
        self.load_options()
        self.load_all_students()
        self.load_deleted_students()
        self.show_main_menu()
    
    def load_options(self):
        """Load school and course options from JSON file"""
        if os.path.exists(self.options_file):
            try:
                with open(self.options_file, 'r') as f:
                    options_data = json.load(f)
                    self.school_options = options_data.get('schools', self.school_options)
                    self.course_mapping = options_data.get('courses', self.course_mapping)
            except:
                # If file is corrupted, use defaults
                pass
    
    def save_options(self):
        """Save school and course options to JSON file"""
        options_data = {
            'schools': self.school_options,
            'courses': self.course_mapping
        }
        with open(self.options_file, 'w') as f:
            json.dump(options_data, f, indent=4)
    
    def load_all_students(self):
        """Load all student data from JSON file"""
        if os.path.exists(self.data_file):
            with open(self.data_file, 'r') as f:
                self.all_students = json.load(f)
        else:
            self.all_students = []
    
    def save_all_students(self):
        """Save all student data to JSON file"""
        with open(self.data_file, 'w') as f:
            json.dump(self.all_students, f, indent=4)
    
    def load_deleted_students(self):
        """Load deleted student data from JSON file (trash)"""
        if os.path.exists(self.trash_file):
            with open(self.trash_file, 'r') as f:
                self.deleted_students = json.load(f)
        else:
            self.deleted_students = []
    
    def save_deleted_students(self):
        """Save deleted student data to JSON file (trash)"""
        with open(self.trash_file, 'w') as f:
            json.dump(self.deleted_students, f, indent=4)
    
    def clear_window(self):
        """Clear all widgets from window"""
        for widget in self.root.winfo_children():
            widget.destroy()
    
    def show_main_menu(self):
        """Display main menu with 4 buttons"""
        self.clear_window()
        
        # Configure root grid weights for responsiveness
        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_columnconfigure(0, weight=1)
        
        title = tk.Label(self.root, text="Student Information System", 
                        font=("Arial", 18, "bold"))
        title.grid(row=0, column=0, pady=20, padx=20, sticky="ew")
        
        # Create frame for buttons with responsive sizing
        button_frame = tk.Frame(self.root)
        button_frame.grid(row=1, column=0, pady=20, padx=20, sticky="nsew")
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_rowconfigure(0, weight=1)
        button_frame.grid_rowconfigure(1, weight=1)
        button_frame.grid_rowconfigure(2, weight=1)
        button_frame.grid_rowconfigure(3, weight=1)
        button_frame.grid_rowconfigure(4, weight=1)
        
        btn_register = tk.Button(button_frame, text="Register Student", 
                                 font=("Arial", 12), padx=20, pady=15,
                                 command=self.show_register_screen)
        btn_register.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        btn_renewal = tk.Button(button_frame, text="Renewal", 
                               font=("Arial", 12), padx=20, pady=15,
                               command=self.show_renewal_screen)
        btn_renewal.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        
        btn_list = tk.Button(button_frame, text="List of Registered Students", 
                            font=("Arial", 12), padx=20, pady=15,
                            command=self.show_list_students)
        btn_list.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        
        btn_export = tk.Button(button_frame, text="Print to Excel", 
                              font=("Arial", 12), padx=20, pady=15,
                              command=self.show_export_options)
        btn_export.grid(row=3, column=0, sticky="nsew", padx=10, pady=10)

        btn_payroll = tk.Button(button_frame, text="Print to Payroll", 
                                font=("Arial", 12), padx=20, pady=15,
                                bg="#6a1b9a", fg="white",
                                activebackground="#4a148c", activeforeground="white",
                                command=self.show_payroll_options)
        btn_payroll.grid(row=4, column=0, sticky="nsew", padx=10, pady=10)
    
    def show_register_screen(self):
        """Display student registration form"""
        self.clear_window()
        
        # Configure grid weights
        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_columnconfigure(0, weight=1)
        
        title = tk.Label(self.root, text="Student Registration", 
                        font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, pady=10, padx=10, sticky="ew")
        
        # Create main frame with scrollbar
        canvas_frame = tk.Frame(self.root)
        canvas_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)
        
        canvas = tk.Canvas(canvas_frame)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Configure scrollable frame for responsive text entry
        self.root.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units") if canvas.winfo_exists() else None)
        self.root.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units") if canvas.winfo_exists() else None)
        self.root.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units") if canvas.winfo_exists() else None)
        
        # Full Name
        tk.Label(scrollable_frame, text="Full Name:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        full_name_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        full_name_entry.pack(pady=5, padx=20, fill="x")
        
        # Barangay
        tk.Label(scrollable_frame, text="Barangay:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        barangay_options = [
            "Abar 1st", "Calaocan", "Canuto Ramos", "Crisanto Sanchez", "Ferdinand E. Marcos",
            "Malasin", "Rafael Rueda Sr.", "Raymundo Eugenio", "Sibut", "Sto. Niño 1st",
            "Abar 2nd", "A. Pascual", "Bagong Sikat", "Caanawan", "Camanacsacan",
            "Culaylay", "Dizol", "Kaliwanagan", "Kita-Kita", "Manicla",
            "Palestina", "Parang Mangga", "Pinili", "Porais", "San Agustin",
            "San Juan", "San Mauricio", "Sto. Niño 2nd", "Sto. Niño 3rd", "Sto. Tomas",
            "Sinipit Bubon", "Tabulac", "Tayabo", "Tondod", "Tulat",
            "Villa Floresta", "Villa Joson", "Villa Marina"
        ]
        barangay_var = tk.StringVar()
        barangay_combo = ttk.Combobox(scrollable_frame, textvariable=barangay_var, values=barangay_options, state="normal")
        barangay_combo.pack(pady=5, padx=20, fill="x")

        def update_barangay_options(event=None):
            text = barangay_var.get().strip()
            lowercase = text.lower()
            filtered = [b for b in barangay_options if lowercase in b.lower()] if text else barangay_options
            barangay_combo['values'] = filtered

            # Strict first-match inline autocomplete
            if text:
                match = next((b for b in barangay_options if b.lower().startswith(lowercase)), None)
                if match:
                    # set full text and highlight suggested part
                    barangay_var.set(match)
                    try:
                        barangay_combo.icursor(len(text))
                        barangay_combo.select_range(len(text), len(match))
                    except Exception:
                        pass

        barangay_combo.bind('<KeyRelease>', update_barangay_options)
        barangay_combo.bind('<FocusIn>', lambda e: update_barangay_options() if not barangay_var.get() else None)
        
        # Address
        tk.Label(scrollable_frame, text="Address:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        address_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        address_entry.pack(pady=5, padx=20, fill="x")
        
        # Contact Number
        tk.Label(scrollable_frame, text="Contact Number:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        contact_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        contact_entry.pack(pady=5, padx=20, fill="x")
        
        # Course mapping for each school
        self.course_mapping = {
            'PHINMA Araullo University San Jose City ': [
                'Accountancy', 'Business Administration', 'Hospitality Management', 'Criminology',
                'Information Technology', 'Computer Science', 'Education (Elementary & Secondary)',
                'Nursing', 'Civil Engineering'
            ],
            'Core Gateway College, Inc.': [
                'Computer Science', 'Business Administration (Management, Accounting)', 'Political Science',
                'Education (Elementary & Secondary)', "Master's: Education, Public Administration"
            ],
            'San Jose Christian Colleges': [
                'Business Administration', 'Education', 'Criminology', 'Information Technology',
                'Hospitality / Tourism'
            ],
            'Golden Success University': [
                'Business Administration', 'Education', 'Information Technology', 'Criminology'
            ],
            'STI College San Jose': [
                'Information Technology', 'Computer Science', 'Hospitality Management',
                'Tourism Management', 'Business Administration'
            ],
            'Central Luzon State University': [
                'BS Agriculture',
                'BS Agribusiness',
                'BS Agricultural and Biosystems Engineering',
                'BS Animal Science / Animal Husbandry',
                'BS Fisheries / Aquaculture',
                'BS Biology',
                'BS Chemistry',
                'BS Environmental Science',
                'BS Civil Engineering',
                'BS Information Technology',
                'BS Accountancy',
                'BS Accounting Technology',
                'BS Business Administration',
                'BS Entrepreneurship',
                'Bachelor of Elementary Education (BEEd)',
                'Bachelor of Secondary Education (BSEd) – English',
                'Bachelor of Secondary Education (BSEd) – Math',
                'Bachelor of Secondary Education (BSEd) – Science',
                'Bachelor of Secondary Education (BSEd) – Filipino',
                'Bachelor of Secondary Education (BSEd) – Social Studies',
                'Bachelor of Secondary Education (BSEd) – MAPEH',
                'Bachelor of Secondary Education (BSEd) – TLE',
                'Bachelor of Early Childhood Education',
                'Bachelor of Physical Education',
                'BA Psychology',
                'BA Social Sciences',
                'BA Filipino',
                'BA Language / Literature',
                'BS Hospitality Management',
                'BS Food Technology',
                'BS Fashion and Textile Technology',
                'BS Development Communication',
                'Doctor of Veterinary Medicine (DVM)'
            ]
        }
        
        # School
        tk.Label(scrollable_frame, text="School:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        school_var = tk.StringVar()
        school_combo = ttk.Combobox(scrollable_frame, textvariable=school_var, values=self.school_options, state="normal", font=("Arial", 10))
        school_combo.pack(pady=5, padx=20, fill="x")
        
        # Course
        tk.Label(scrollable_frame, text="Course:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        course_var = tk.StringVar()
        course_combo = ttk.Combobox(scrollable_frame, textvariable=course_var, state="normal", font=("Arial", 10))
        course_combo.pack(pady=5, padx=20, fill="x")
        
        # Bind school selection to update course options
        def update_course_options(event=None):
            selected_school = school_var.get()
            courses = self.course_mapping.get(selected_school, [])
            course_combo['values'] = courses
            course_combo.set('')

        school_combo.bind('<<ComboboxSelected>>', update_course_options)

        course_typeahead_id = None

        def apply_course_typeahead(text, lowercase, base_courses, filtered):
            if len(text) < 2 or not filtered:
                return
            match = next((c for c in base_courses if c.lower().startswith(lowercase)), None)
            if match:
                course_var.set(match)
                try:
                    course_combo.icursor(len(text))
                    course_combo.select_range(len(text), len(match))
                except Exception:
                    pass

        def update_school_options(event=None):
            text = school_var.get().strip()
            lowercase = text.lower()
            filtered = [s for s in self.school_options if lowercase in s.lower()] if text else self.school_options
            school_combo['values'] = filtered

            # Handle "Others" selection - allow manual input
            if text == "Others":
                school_combo['state'] = 'normal'  # Allow typing
                course_combo['state'] = 'normal'  # Allow typing for courses too
                course_combo['values'] = []  # Clear course options for manual input
                return

            # Copy Barangay-style first match inline autocomplete (no delayed typeahead)
            if text and filtered:
                match = next((s for s in self.school_options if s.lower().startswith(lowercase)), None)
                if match and match != "Others":
                    school_var.set(match)
                    try:
                        school_combo.icursor(len(text))
                        school_combo.select_range(len(text), len(match))
                    except Exception:
                        pass
                    update_course_options()

            if filtered:
                school_combo.event_generate('<Down>')

        school_combo.bind('<KeyRelease>', update_school_options)

        def update_course_filter(event=None):
            nonlocal course_typeahead_id
            selected_school = school_var.get()
            
            # If "Others" is selected for school, allow manual course input
            if selected_school == "Others":
                course_combo['values'] = []
                return
                
            base_courses = self.course_mapping.get(selected_school, [])
            text = course_var.get().strip()
            lowercase = text.lower()
            filtered = [c for c in base_courses if lowercase in c.lower()] if text else base_courses
            course_combo['values'] = filtered

            if course_typeahead_id:
                self.root.after_cancel(course_typeahead_id)
                course_typeahead_id = None
            if len(text) >= 2 and filtered:
                course_typeahead_id = self.root.after(200, lambda: apply_course_typeahead(text, lowercase, base_courses, filtered))

            if filtered:
                course_combo.event_generate('<Down>')

        course_combo.bind('<KeyRelease>', update_course_filter)
        
        # Year Level
        tk.Label(scrollable_frame, text="Year Level:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        school_year_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        school_year_entry.pack(pady=5, padx=20, fill="x")
        
        # Batch/Year Level
        tk.Label(scrollable_frame, text="Batch (1-7):", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        batch_combo = ttk.Combobox(scrollable_frame, values=['1', '2', '3', '4', '5', '6', '7'], state="readonly")
        batch_combo.pack(pady=5, padx=20, fill="x")
        
        # Status
        tk.Label(scrollable_frame, text="Status:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        status_label = tk.Label(scrollable_frame, text="Status will be determined automatically based on documents", 
                               font=("Arial", 9), fg="blue", anchor="w")
        status_label.pack(pady=5, padx=40, fill="x")
        
        # Documents checkboxes
        tk.Label(scrollable_frame, text="Required Documents:", font=("Arial", 10, "bold")).pack(pady=10, padx=20, fill="x")
        
        # Select All checkbox
        select_all_var = tk.BooleanVar()
        select_all_chk = tk.Checkbutton(scrollable_frame, text="Select All Documents", variable=select_all_var, 
                                       font=("Arial", 10, "bold"), anchor="w")
        select_all_chk.pack(anchor="w", padx=20, pady=5, fill="x")
        
        doc_vars = {}
        documents = [
            'Certificate of Residency',
            'Pagpapatunay Form',
            'Picture of the House',
            'Good Moral Certificate',
            'Original Certificate of Grades',
            'Proof of Enrollment',
            'School ID'
        ]
        
        def toggle_all_documents():
            """Toggle all document checkboxes based on Select All state"""
            state = select_all_var.get()
            for var in doc_vars.values():
                var.set(state)
        
        # Bind the select all checkbox to toggle function
        select_all_var.trace_add('write', lambda *args: toggle_all_documents())
        
        for doc in documents:
            var = tk.BooleanVar()
            doc_vars[doc] = var
            chk = tk.Checkbutton(scrollable_frame, text=doc, variable=var, font=("Arial", 10), anchor="w")
            chk.pack(anchor="w", padx=40, pady=3, fill="x")
        
        # Button frame
        button_frame = tk.Frame(scrollable_frame)
        button_frame.pack(pady=20, fill="x", padx=20)
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        
        def save_register():
            name = full_name_entry.get().strip()
            if not name:
                messagebox.showerror("Error", "Please enter full name")
                return

            # Prevent duplicate student registration by name (case-insensitive)
            existing_names = [student.get('full_name', '').strip().lower() for student in self.all_students]
            if name.lower() in existing_names:
                messagebox.showerror("Duplicate Entry", "That student is already registered.")
                return
            
            school_name = school_combo.get().strip()
            course_name = course_combo.get().strip()
            
            # Add new school to options if entered manually
            if school_name and school_name not in self.school_options:
                self.school_options.append(school_name)
                self.course_mapping[school_name] = []
            
            # Add new course to mapping if entered manually
            if course_name and school_name:
                if course_name not in self.course_mapping.get(school_name, []):
                    self.course_mapping[school_name].append(course_name)
            
            # Save updated options
            self.save_options()
            
            # Check if all required documents are checked
            all_documents_checked = all(var.get() for var in doc_vars.values())
            status = "Complete" if all_documents_checked else "Incomplete"
            
            data = {
                'full_name': name,
                'barangay': barangay_combo.get(),
                'address': address_entry.get(),
                'contact_number': contact_entry.get(),
                'school': school_name,
                'course': course_name,
                'school_year': school_year_entry.get(),
                'batch': batch_combo.get(),
                'status': status,
                'documents': {doc: var.get() for doc, var in doc_vars.items()},
                'registration_date': datetime.now().strftime("%Y-%m-%d")
            }
            
            self.all_students.append(data)
            self.save_all_students()
            messagebox.showinfo("Success", f"Student registered successfully! Status: {status}")
            self.show_main_menu()
        
        save_btn = tk.Button(button_frame, text="Save Student", command=save_register, padx=10, pady=8)
        save_btn.pack(side="left", padx=5, fill="x", expand=True)
        
        back_btn = tk.Button(button_frame, text="Back", command=self.show_main_menu, padx=10, pady=8)
        back_btn.pack(side="left", padx=5, fill="x", expand=True)
    
    def show_renewal_screen(self):
        """Display renewal form with requirements"""
        self.clear_window()
        
        # Configure grid weights
        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_columnconfigure(0, weight=1)
        
        title = tk.Label(self.root, text="Student Renewal", font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, pady=10, padx=10, sticky="ew")
        
        # Center frame for content
        center_frame = tk.Frame(self.root)
        center_frame.grid(row=1, column=0, pady=20, padx=20, sticky="n")
        center_frame.grid_columnconfigure(0, weight=1)
        
        # Student selection
        tk.Label(center_frame, text="Select Student:", font=("Arial", 10)).pack(pady=5, padx=20, fill="x")
        student_names = [s['full_name'] for s in self.all_students]
        
        if not student_names:
            messagebox.showwarning("No Data", "No registered students found. Please register a student first.")
            self.show_main_menu()
            return
        
        def update_student_options(event=None):
            typed = student_var.get().strip().lower()
            if typed == "":
                filtered = student_names
            else:
                filtered = [name for name in student_names if typed in name.lower()]
            student_combo['values'] = filtered
            # Keep user typing and optionally drop down
            if filtered:
                student_combo.event_generate('<Down>')

        student_var = tk.StringVar()
        student_combo = ttk.Combobox(center_frame, textvariable=student_var, 
                                     values=student_names, state="normal")
        student_combo.pack(pady=5, padx=20, fill="x")
        student_combo.bind('<KeyRelease>', update_student_options)
        student_combo.bind('<<ComboboxSelected>>', lambda e: update_student_options())
        
        # Renewal requirements
        tk.Label(center_frame, text="Renewal Requirements:", font=("Arial", 12, "bold")).pack(pady=10, padx=20, fill="x")
        
        req_vars = {}
        requirements = [
            'Last grades from last semester',
            'Liquidation of 5000 from LGU',
            'Certification of enrollment'
        ]
        
        for req in requirements:
            var = tk.BooleanVar()
            req_vars[req] = var
            chk = tk.Checkbutton(center_frame, text=req, variable=var, font=("Arial", 10))
            chk.pack(anchor="w", padx=40, pady=5, fill="x")
        
        # Button frame
        button_frame = tk.Frame(center_frame)
        button_frame.pack(pady=20, fill="x", padx=20)
        
        def save_renewal():
            if not student_var.get():
                messagebox.showerror("Error", "Please select a student")
                return
            
            # Check if all requirements are met
            if not all(req_vars.values()):
                messagebox.showerror("Error", "All renewal requirements must be checked")
                return
            
            # Find and update student
            for student in self.all_students:
                if student['full_name'] == student_var.get():
                    student['renewal_date'] = datetime.now().strftime("%Y-%m-%d")
                    student['renewal_requirements'] = {req: var.get() for req, var in req_vars.items()}
                    break
            
            self.save_all_students()
            messagebox.showinfo("Success", "Student renewed successfully!")
            self.show_main_menu()
        
        save_btn = tk.Button(button_frame, text="Save Renewal", command=save_renewal, padx=10, pady=8)
        save_btn.pack(side="left", padx=5, fill="x", expand=True)
        
        back_btn = tk.Button(button_frame, text="Back", command=self.show_main_menu, padx=10, pady=8)
        back_btn.pack(side="left", padx=5, fill="x", expand=True)
    
    def create_student_tab(self, notebook, tab_name, students, description):
        """Create a tab with student list for the notebook"""
        tab_frame = tk.Frame(notebook)
        notebook.add(tab_frame, text=tab_name)
        
        # Description label
        desc_label = tk.Label(tab_frame, text=description, font=("Arial", 10, "italic"), fg="gray")
        desc_label.pack(pady=(10, 5), padx=10, anchor="w")
        
        if not students:
            no_data_label = tk.Label(tab_frame, text=f"No {tab_name.lower()} found", font=("Arial", 12))
            no_data_label.pack(pady=20)
            return None

        # Create frame for treeview and scrollbar
        table_frame = tk.Frame(tab_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Create treeview
        columns = ('Name', 'Barangay', 'School', 'Course', 'School Year', 'Batch', 'Status', 'Date')
        tree = ttk.Treeview(table_frame, columns=columns, height=12)
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        h_scrollbar = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        tree.column('#0', width=0, stretch=tk.NO)
        tree.column('Name', anchor=tk.W, width=120, minwidth=80)
        tree.column('Barangay', anchor=tk.W, width=100, minwidth=60)
        tree.column('School', anchor=tk.W, width=100, minwidth=60)
        tree.column('Course', anchor=tk.W, width=80, minwidth=50)
        tree.column('School Year', anchor=tk.CENTER, width=90, minwidth=60)
        tree.column('Batch', anchor=tk.CENTER, width=50, minwidth=40)
        tree.column('Status', anchor=tk.CENTER, width=70, minwidth=50)
        tree.column('Date', anchor=tk.CENTER, width=90, minwidth=70)
        
        tree.heading('#0', text='', anchor=tk.W)
        tree.heading('Name', text='Name', anchor=tk.W)
        tree.heading('Barangay', text='Barangay', anchor=tk.W)
        tree.heading('School', text='School', anchor=tk.W)
        tree.heading('Course', text='Course', anchor=tk.W)
        tree.heading('School Year', text='School Year', anchor=tk.CENTER)
        tree.heading('Batch', text='Batch', anchor=tk.CENTER)
        tree.heading('Status', text='Status', anchor=tk.CENTER)
        tree.heading('Date', text='Date', anchor=tk.CENTER)
        
        for student in students:
            tree.insert(parent='', index='end', iid=student['full_name'],
                       text='',
                       values=(student['full_name'], student.get('barangay', ''), student['school'], 
                              student['course'], student.get('school_year', ''), student.get('batch', ''),
                              student.get('status', ''), student.get('registration_date', '')))
        
        # Grid layout for scrollbars
        tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # Bind double-click to show student details
        tree.bind("<Double-1>", lambda event: self.show_student_details(tree, students))

        return tree
    
    def show_list_students(self):
        """Display list of all registered students with categorization"""
        self.clear_window()
        
        # Configure grid weights
        self.root.grid_rowconfigure(0, weight=0)  # Title
        self.root.grid_rowconfigure(1, weight=0)  # Filters
        self.root.grid_rowconfigure(2, weight=0)  # Instruction
        self.root.grid_rowconfigure(3, weight=1)  # Content
        self.root.grid_rowconfigure(4, weight=0)  # Summary
        self.root.grid_rowconfigure(5, weight=0)  # Buttons
        self.root.grid_columnconfigure(0, weight=1)
        
        title = tk.Label(self.root, text="Registered Students", font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, pady=10, padx=10, sticky="ew")
        
        # Filter frame
        filter_frame = tk.Frame(self.root, bg="#f5f5f5", relief="groove", bd=1)
        filter_frame.grid(row=1, column=0, pady=(0, 10), padx=10, sticky="ew")
        filter_frame.grid_columnconfigure(0, weight=1)
        filter_frame.grid_columnconfigure(1, weight=0)
        filter_frame.grid_columnconfigure(2, weight=0)
        filter_frame.grid_columnconfigure(3, weight=0)
        
        # Search bar for names
        search_label = tk.Label(filter_frame, text="Search Name:", font=("Arial", 10, "bold"), bg="#f5f5f5")
        search_label.grid(row=0, column=0, padx=(10, 5), pady=10, sticky="w")
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(filter_frame, textvariable=search_var, font=("Arial", 10), width=30)
        search_entry.grid(row=0, column=1, padx=(0, 10), pady=10, sticky="ew")
        
        # Batch filter
        batch_label = tk.Label(filter_frame, text="Batch:", font=("Arial", 10, "bold"), bg="#f5f5f5")
        batch_label.grid(row=0, column=2, padx=(0, 5), pady=10, sticky="w")
        
        batch_var = tk.StringVar(value="All")
        batch_combo = ttk.Combobox(filter_frame, textvariable=batch_var, 
                                  values=["All", "1", "2", "3", "4", "5", "6", "7"], 
                                  state="readonly", width=8)
        batch_combo.grid(row=0, column=3, padx=(0, 10), pady=10, sticky="w")
        
        # Year level filter
        year_label = tk.Label(filter_frame, text="Year Level:", font=("Arial", 10, "bold"), bg="#f5f5f5")
        year_label.grid(row=0, column=4, padx=(0, 5), pady=10, sticky="w")
        
        year_var = tk.StringVar()
        year_entry = tk.Entry(filter_frame, textvariable=year_var, font=("Arial", 10), width=15)
        year_entry.grid(row=0, column=5, padx=(0, 10), pady=10, sticky="w")
        
        # Clear filters button
        clear_btn = tk.Button(filter_frame, text="Clear Filters", font=("Arial", 9), 
                             command=lambda: self.clear_filters(search_var, batch_var, year_var))
        clear_btn.grid(row=0, column=6, padx=(0, 10), pady=10)
        
        # Instruction label
        instruction_label = tk.Label(self.root, text="💡 Double-click on any student name to view detailed information", 
                                   font=("Arial", 10, "italic"), fg="blue")
        instruction_label.grid(row=2, column=0, pady=(0, 10), padx=10, sticky="ew")
        
        # Function to filter students
        def filter_students():
            search_text = search_var.get().strip().lower()
            batch_filter = batch_var.get()
            year_filter = year_var.get().strip().lower()
            
            filtered_students = []
            for student in self.all_students:
                # Name search filter
                name_match = search_text == "" or search_text in student.get('full_name', '').lower()
                
                # Batch filter
                batch_match = batch_filter == "All" or student.get('batch', '') == batch_filter
                
                # Year level filter
                year_match = year_filter == "" or year_filter in student.get('school_year', '').lower()
                
                if name_match and batch_match and year_match:
                    filtered_students.append(student)
            
            # Sort students: by batch number first, then alphabetically by name
            def sort_key(student):
                name = student.get('full_name', '').strip().lower()
                batch = student.get('batch', '')
                # Convert batch to int for proper numeric sorting, use 0 for empty/invalid batches
                try:
                    batch_num = int(batch) if batch.isdigit() else 0
                except:
                    batch_num = 0
                return (batch_num, name)
            
            filtered_students.sort(key=sort_key)
            
            return filtered_students
        
        # Function to update display
        def update_display():
            filtered_students = filter_students()
            
            # Clear existing notebook
            for widget in self.root.grid_slaves(row=3, column=0):
                widget.destroy()
            
            if not filtered_students:
                no_results_label = tk.Label(self.root, text="No students match the current filters", 
                                          font=("Arial", 12), fg="red")
                no_results_label.grid(row=3, column=0, pady=20)
                
                # Update summary
                summary_frame = tk.Frame(self.root)
                summary_frame.grid(row=4, column=0, pady=5, padx=10, sticky="ew")
                summary_label = tk.Label(summary_frame, text="No students found", 
                                       font=("Arial", 10, "italic"), fg="red")
                summary_label.pack()
            else:
                # Create notebook for tabs
                notebook = ttk.Notebook(self.root)
                notebook.grid(row=3, column=0, sticky="nsew", padx=10, pady=10)
                
                # Separate filtered students by renewal status
                new_students = [s for s in filtered_students if 'renewal_date' not in s]
                renewed_students = [s for s in filtered_students if 'renewal_date' in s]
                
                # Create tabs
                self.create_student_tab(notebook, "All Students", filtered_students, "All registered students")
                self.create_student_tab(notebook, "New Students", new_students, "Students who have registered but not yet renewed")
                self.create_student_tab(notebook, "Renewed Students", renewed_students, "Students who have completed the renewal process")
                
                # Update summary
                summary_frame = tk.Frame(self.root)
                summary_frame.grid(row=4, column=0, pady=5, padx=10, sticky="ew")
                summary_text = f"Filtered Results: {len(filtered_students)} students | New: {len(new_students)} | Renewed: {len(renewed_students)}"
                summary_label = tk.Label(summary_frame, text=summary_text, font=("Arial", 10, "italic"), fg="green")
                summary_label.pack()
        
        # Bind filter updates
        search_entry.bind("<KeyRelease>", lambda e: update_display())
        batch_combo.bind("<<ComboboxSelected>>", lambda e: update_display())
        year_entry.bind("<KeyRelease>", lambda e: update_display())
        
        # Initial display
        update_display()
        
        # Back button
        button_frame = tk.Frame(self.root)
        button_frame.grid(row=5, column=0, pady=10)
        trash_btn = tk.Button(button_frame, text=f"Trash ({len(self.deleted_students)})", command=self.show_trash, bg="#ff9800", fg="white", padx=20, pady=8)
        trash_btn.pack(side="left", padx=5)

        back_btn = tk.Button(button_frame, text="Back", command=self.show_main_menu, padx=20, pady=8)
        back_btn.pack(side="left", padx=5)
    
    def clear_filters(self, search_var, batch_var, year_var):
        """Clear all filter values"""
        search_var.set("")
        batch_var.set("All")
        year_var.set("")

    def show_export_options(self):
        """Show buttons for export filters: All, Unrenewed, Renewed"""
        self.clear_window()

        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_columnconfigure(0, weight=1)

        title = tk.Label(self.root, text="Export Students to Excel", font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, pady=15, padx=15, sticky="ew")

        button_frame = tk.Frame(self.root)
        button_frame.grid(row=1, column=0, pady=10, padx=20, sticky="nsew")
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_rowconfigure(0, weight=1)
        button_frame.grid_rowconfigure(1, weight=1)
        button_frame.grid_rowconfigure(2, weight=1)

        btn_all = tk.Button(button_frame, text="All Students", font=("Arial", 12), padx=20, pady=15,
                            bg="#4caf50", fg="white",
                            activebackground="#45a049", activeforeground="white",
                            command=lambda: self.export_to_excel("all"))
        btn_all.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        btn_unrenewed = tk.Button(button_frame, text="Unrenewed Students", font=("Arial", 12), padx=20, pady=15,
                                  bg="#f44336", fg="white",
                                  activebackground="#e53935", activeforeground="white",
                                  command=lambda: self.export_to_excel("unrenewed"))
        btn_unrenewed.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        btn_renewed = tk.Button(button_frame, text="Renewed Students", font=("Arial", 12), padx=20, pady=15,
                                bg="#2196f3", fg="white",
                                activebackground="#1e88e5", activeforeground="white",
                                command=lambda: self.export_to_excel("renewed"))
        btn_renewed.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)

        back_btn = tk.Button(self.root, text="Back", font=("Arial", 11), padx=20, pady=10,
                             bg="#757575", fg="white", activebackground="#616161", activeforeground="white",
                             command=self.show_main_menu)
        back_btn.grid(row=2, column=0, pady=10)

    def show_payroll_options(self):
        """Show payroll print options by renewal status and batch."""
        self.clear_window()

        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=0)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_rowconfigure(3, weight=0)
        self.root.grid_rowconfigure(4, weight=0)
        self.root.grid_columnconfigure(0, weight=1)

        title = tk.Label(self.root, text="Print to Payroll", font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, pady=15, padx=15, sticky="ew")  

        self.payroll_filter_var = tk.StringVar(value="all")
        self.payroll_batch_var = tk.StringVar(value="All")

        radio_frame = tk.Frame(self.root)
        radio_frame.grid(row=1, column=0, pady=10, padx=20, sticky="ew")
        radio_frame.grid_columnconfigure(0, weight=1)

        tk.Radiobutton(radio_frame, text="All Registered", variable=self.payroll_filter_var, value="all", font=("Arial", 12)).pack(fill="x", pady=5)
        tk.Radiobutton(radio_frame, text="Renewed", variable=self.payroll_filter_var, value="renewed", font=("Arial", 12)).pack(fill="x", pady=5)
        tk.Radiobutton(radio_frame, text="Not Renewed", variable=self.payroll_filter_var, value="unrenewed", font=("Arial", 12)).pack(fill="x", pady=5)

        batch_frame = tk.Frame(self.root)
        batch_frame.grid(row=2, column=0, pady=(0, 10), padx=20, sticky="ew")
        batch_frame.grid_columnconfigure(1, weight=1)

        tk.Label(batch_frame, text="Batch:", font=("Arial", 11, "bold")).grid(row=0, column=0, padx=(0, 10), sticky="w")
        batch_combo = ttk.Combobox(
            batch_frame,
            textvariable=self.payroll_batch_var,
            values=["All", "1", "2", "3", "4", "5", "6", "7"],
            state="readonly",
            font=("Arial", 11)
        )
        batch_combo.grid(row=0, column=1, sticky="ew")

        btn_export_payroll = tk.Button(self.root, text="Export Payroll", font=("Arial", 12), bg="#6a1b9a", fg="white",
                                       activebackground="#4a148c", activeforeground="white", padx=20, pady=15,
                                       command=self.print_to_payroll)
        btn_export_payroll.grid(row=3, column=0, pady=10, padx=20, sticky="ew")

        back_btn = tk.Button(self.root, text="Back", font=("Arial", 11), padx=20, pady=10,
                             bg="#757575", fg="white", activebackground="#616161", activeforeground="white",
                             command=self.show_main_menu)
        back_btn.grid(row=4, column=0, pady=10)

    def show_trash(self):
        """Display list of trashed students with restore/permanent-delete actions"""
        self.clear_window()

        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=0)
        self.root.grid_rowconfigure(2, weight=1)
        self.root.grid_rowconfigure(3, weight=0)
        self.root.grid_columnconfigure(0, weight=1)

        title = tk.Label(self.root, text="Trash - Deleted Students", font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, pady=10, padx=10, sticky="ew")

        self.trash_tree = None

        if not self.deleted_students:
            tk.Label(self.root, text="Trash is empty", font=("Arial", 12)).grid(row=1, column=0, pady=20)
        else:
            # Treeview using existing helper
            notebook = ttk.Notebook(self.root)
            notebook.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
            self.trash_tree = self.create_student_tab(notebook, "Deleted", self.deleted_students, "Students in trash")

        action_frame = tk.Frame(self.root)
        action_frame.grid(row=3, column=0, pady=10)

        restore_btn = tk.Button(action_frame, text="Restore Selected", bg="#4CAF50", fg="white", command=self.restore_selected, padx=20, pady=8)
        restore_btn.pack(side="left", padx=5)

        perm_delete_btn = tk.Button(action_frame, text="Delete Permanently", bg="#f44336", fg="white", command=self.permanent_delete_selected, padx=20, pady=8)
        perm_delete_btn.pack(side="left", padx=5)

        back_btn = tk.Button(action_frame, text="Back", command=self.show_list_students, padx=20, pady=8)
        back_btn.pack(side="left", padx=5)

    def get_selected_trash_student(self):
        # helper for restore/permanent-delete
        # Prefer direct trash tree reference for reliability
        tree = getattr(self, 'trash_tree', None)

        if tree is None or not isinstance(tree, ttk.Treeview):
            # fallback: search for first Treeview in current root children
            for widget in self.root.winfo_children():
                if isinstance(widget, ttk.Notebook):
                    tab = widget.nametowidget(widget.select())
                    for child in tab.winfo_children():
                        if isinstance(child, ttk.Treeview):
                            tree = child
                            break
                    break

        if tree is None:
            messagebox.showwarning("No Selection", "Trash table not found")
            return None

        selected = tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a trashed student")
            return None

        student_name = selected[0]
        student = next((s for s in self.deleted_students if s.get('full_name') == student_name), None)
        if not student:
            messagebox.showerror("Error", "Selected student not found in trash list")

        return student

    def restore_selected(self):
        student = self.get_selected_trash_student()
        if not student:
            return

        if messagebox.askyesno("Confirm Restore", f"Restore student '{student['full_name']}'?"):
            self.deleted_students.remove(student)
            self.all_students.append(student)
            self.save_deleted_students()
            self.save_all_students()
            messagebox.showinfo("Success", f"Student '{student['full_name']}' restored")
            self.show_trash()

    def permanent_delete_selected(self):
        student = self.get_selected_trash_student()
        if not student:
            return

        if messagebox.askyesno("Confirm Permanent Delete", f"Permanently delete student '{student['full_name']}'? This cannot be undone."):
            self.deleted_students.remove(student)
            self.save_deleted_students()
            messagebox.showinfo("Deleted", f"Student '{student['full_name']}' has been permanently deleted")
            self.show_trash()
    
    def show_student_details(self, tree, students):
        """Show detailed view of selected student"""
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a student to view details")
            return
        
        student_name = selected_item[0]
        student = next((s for s in students if s['full_name'] == student_name), None)
        if not student:
            messagebox.showerror("Error", "Student not found")
            return
        
        self.show_student_details_from_data(student)
    
    def edit_student_details(self, student):
        """Edit student information from details view"""
        # Create edit dialog
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Student Information")
        edit_window.geometry("500x600")
        edit_window.resizable(False, False)
        
        # Center the window
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # Title
        title_label = tk.Label(edit_window, text="Edit Student Information", font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # Scrollable frame
        canvas = tk.Canvas(edit_window)
        scrollbar = ttk.Scrollbar(edit_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Form fields with current values
        tk.Label(scrollable_frame, text="Full Name:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        name_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        name_entry.insert(0, student.get('full_name', ''))
        name_entry.pack(pady=5, padx=20, fill="x")
        
        # Barangay
        tk.Label(scrollable_frame, text="Barangay:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        barangay_combo = ttk.Combobox(scrollable_frame, values=[
            'A. Bonifacio', 'Abianan', 'Aguinaldo', 'Bagong Sikat', 'Bayanihan', 'Cristo Rey', 'Del Pilar', 'Fatima',
            'Gumamela', 'Holy Child', 'Holy Rosary', 'Immaculate Conception', 'Kalawaan', 'Libertad', 'Little Baguio',
            'Malinis', 'Maluya', 'Mapalad', 'San Antonio', 'San Isidro', 'San Jose', 'San Luis', 'San Miguel',
            'San Rafael', 'San Roque', 'Santa Cruz', 'Santa Lucia', 'Santa Maria', 'Santo Cristo', 'Santo Niño',
            'Santo Rosario', 'Santo Tomas', 'Tibagan', 'Tinigaw', 'Tulay Na Patak', 'Villa Esperanza', 'Villa Jose',
            'Villa Maria'
        ], state="readonly", font=("Arial", 10))
        barangay_combo.set(student.get('barangay', ''))
        barangay_combo.pack(pady=5, padx=20, fill="x")
        
        # Address
        tk.Label(scrollable_frame, text="Address:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        address_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        address_entry.insert(0, student.get('address', ''))
        address_entry.pack(pady=5, padx=20, fill="x")
        
        # Contact Number
        tk.Label(scrollable_frame, text="Contact Number:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        contact_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        contact_entry.insert(0, student.get('contact_number', ''))
        contact_entry.pack(pady=5, padx=20, fill="x")
        
        # School
        tk.Label(scrollable_frame, text="School:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        school_combo = ttk.Combobox(scrollable_frame, values=self.school_options, state="normal", font=("Arial", 10))
        school_combo.set(student.get('school', ''))
        school_combo.pack(pady=5, padx=20, fill="x")
        
        # Course
        tk.Label(scrollable_frame, text="Course:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        course_combo = ttk.Combobox(scrollable_frame, state="normal", font=("Arial", 10))
        course_combo.pack(pady=5, padx=20, fill="x")
        
        # Set initial course options and value based on selected school
        selected_school = student.get('school', '')
        if selected_school in self.course_mapping:
            course_combo['values'] = self.course_mapping[selected_school]
        course_combo.set(student.get('course', ''))
        
        # Bind school selection to update course options
        def update_course_options(event):
            selected_school = school_combo.get()
            if selected_school in self.course_mapping:
                course_combo['values'] = self.course_mapping[selected_school]
                course_combo.set('')  # Clear current selection
            else:
                course_combo['values'] = []
                course_combo.set('')
        
        school_combo.bind('<<ComboboxSelected>>', update_course_options)
        
        # School Year
        tk.Label(scrollable_frame, text="School Year:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        school_year_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        school_year_entry.insert(0, student.get('school_year', ''))
        school_year_entry.pack(pady=5, padx=20, fill="x")
        
        # Batch
        tk.Label(scrollable_frame, text="Batch (1-7):", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        batch_combo = ttk.Combobox(scrollable_frame, values=['1', '2', '3', '4', '5', '6', '7'], state="readonly")
        batch_combo.set(student.get('batch', ''))
        batch_combo.pack(pady=5, padx=20, fill="x")
        
        # Documents section
        tk.Label(scrollable_frame, text="Documents Submitted:", font=("Arial", 12, "bold")).pack(pady=(20, 10), padx=20, anchor="w")
        
        documents = [
            'Certificate of Residency',
            'Pagpapatunay Form',
            'Picture of the House',
            'Good Moral Certificate',
            'Original Certificate of Grades',
            'Proof of Enrollment',
            'School ID'
        ]
        
        doc_vars = {}
        for doc in documents:
            var = tk.BooleanVar(value=student.get('documents', {}).get(doc, False))
            doc_vars[doc] = var
            chk = tk.Checkbutton(scrollable_frame, text=doc, variable=var, font=("Arial", 10), anchor="w")
            chk.pack(anchor="w", padx=40, pady=3, fill="x")
        
        # Button frame
        button_frame = tk.Frame(scrollable_frame)
        button_frame.pack(pady=20, fill="x", padx=20)
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        
        def save_changes():
            # Validate required fields
            if not name_entry.get().strip():
                messagebox.showerror("Error", "Full name is required")
                return
            
            school_name = school_combo.get().strip()
            course_name = course_combo.get().strip()
            
            # Add new school to options if entered manually
            if school_name and school_name not in self.school_options:
                self.school_options.append(school_name)
                self.course_mapping[school_name] = []
            
            # Add new course to mapping if entered manually
            if course_name and school_name:
                if course_name not in self.course_mapping.get(school_name, []):
                    self.course_mapping[school_name].append(course_name)
            
            # Save updated options
            self.save_options()
            
            # Update student data
            student['full_name'] = name_entry.get().strip()
            student['barangay'] = barangay_combo.get()
            student['address'] = address_entry.get().strip()
            student['contact_number'] = contact_entry.get().strip()
            student['school'] = school_name
            student['course'] = course_name
            student['school_year'] = school_year_entry.get().strip()
            student['batch'] = batch_combo.get()
            
            # Update documents
            student['documents'] = {doc: var.get() for doc, var in doc_vars.items()}
            
            # Recalculate status
            submitted_docs = sum(var.get() for var in doc_vars.values())
            student['status'] = 'Completed' if submitted_docs == len(documents) else f'Incomplete ({submitted_docs}/{len(documents)})'
            
            # Save to file
            self.save_all_students()
            
            messagebox.showinfo("Success", "Student information updated successfully!")
            edit_window.destroy()
            
            # Refresh the details view
            self.show_student_details_from_student(student)
        
        def cancel_edit():
            edit_window.destroy()
        
        # Save and Cancel buttons
        save_btn = tk.Button(button_frame, text="Save Changes", command=save_changes, 
                           bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), padx=20)
        save_btn.grid(row=0, column=0, padx=5, pady=5)
        
        cancel_btn = tk.Button(button_frame, text="Cancel", command=cancel_edit, 
                             bg="#f44336", fg="white", font=("Arial", 10, "bold"), padx=20)
        cancel_btn.grid(row=0, column=1, padx=5, pady=5)
    
    def show_student_details_from_student(self, student):
        """Show student details after editing"""
        # Find the student in all_students to get updated data
        current_student = next((s for s in self.all_students if s['full_name'] == student['full_name']), student)
        self.show_student_details_from_data(current_student)
    
    def show_student_details_from_data(self, student):
        """Show detailed view of student from data"""
        # Clear window and show student details
        self.clear_window()
        
        # Configure grid weights
        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_rowconfigure(2, weight=0)
        self.root.grid_columnconfigure(0, weight=1)
        
        # Title
        title = tk.Label(self.root, text=f"Student Details - {student['full_name']}", 
                        font=("Arial", 16, "bold"))
        title.grid(row=0, column=0, pady=10, padx=10, sticky="ew")
        
        # Main content frame
        content_frame = tk.Frame(self.root)
        content_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        content_frame.grid_columnconfigure(0, weight=1)
        
        # Create scrollable frame for details
        canvas = tk.Canvas(content_frame)
        scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Student Information Section
        info_frame = tk.Frame(scrollable_frame, relief="groove", bd=2)
        info_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(info_frame, text="Student Information", font=("Arial", 14, "bold"), 
                fg="blue").pack(pady=(10, 15))
        
        # Create info labels
        info_labels = [
            ("Full Name:", student.get('full_name', 'N/A')),
            ("Barangay:", student.get('barangay', 'N/A')),
            ("Address:", student.get('address', 'N/A')),
            ("Contact Number:", student.get('contact_number', 'N/A')),
            ("School:", student.get('school', 'N/A')),
            ("Course:", student.get('course', 'N/A')),
            ("School Year:", student.get('school_year', 'N/A')),
            ("Batch:", student.get('batch', 'N/A')),
            ("Status:", student.get('status', 'N/A')),
            ("Registration Date:", student.get('registration_date', 'N/A')),
        ]
        
        if 'renewal_date' in student:
            info_labels.append(("Renewal Date:", student.get('renewal_date', 'N/A')))
        
        for label_text, value in info_labels:
            row_frame = tk.Frame(info_frame)
            row_frame.pack(fill="x", padx=20, pady=2)
            
            tk.Label(row_frame, text=label_text, font=("Arial", 11, "bold"), 
                    width=15, anchor="w").pack(side="left")
            tk.Label(row_frame, text=str(value), font=("Arial", 11), 
                    anchor="w").pack(side="left", fill="x", expand=True)
        
        # Documents Section
        docs_frame = tk.Frame(scrollable_frame, relief="groove", bd=2)
        docs_frame.pack(fill="x", padx=10, pady=(20, 10))
        
        tk.Label(docs_frame, text="Submitted Documents", font=("Arial", 14, "bold"), 
                fg="green").pack(pady=(10, 15))
        
        documents = [
            'Certificate of Residency',
            'Pagpapatunay Form', 
            'Picture of the House',
            'Good Moral Certificate',
            'Original Certificate of Grades',
            'Proof of Enrollment',
            'School ID'
        ]
        
        submitted_docs = student.get('documents', {})
        
        for doc in documents:
            status = "✓ Submitted" if submitted_docs.get(doc, False) else "✗ Not Submitted"
            color = "green" if submitted_docs.get(doc, False) else "red"
            
            doc_frame = tk.Frame(docs_frame)
            doc_frame.pack(fill="x", padx=20, pady=2)
            
            tk.Label(doc_frame, text=doc, font=("Arial", 10), anchor="w").pack(side="left", fill="x", expand=True)
            tk.Label(doc_frame, text=status, font=("Arial", 10, "bold"), fg=color).pack(side="right")
        
        # Button frame at bottom
        button_frame = tk.Frame(self.root)
        button_frame.grid(row=2, column=0, pady=20, padx=20, sticky="ew")
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        button_frame.grid_columnconfigure(2, weight=1)
        
        # Return button
        return_btn = tk.Button(button_frame, text="Return to List", 
                             command=self.show_list_students,
                             bg="#2196F3", fg="white", font=("Arial", 11, "bold"), padx=20, pady=8)
        return_btn.grid(row=0, column=0, padx=5)
        
        # Edit button
        edit_btn = tk.Button(button_frame, text="Edit Student", 
                           command=lambda: self.edit_student_details(student),
                           bg="#4CAF50", fg="white", font=("Arial", 11, "bold"), padx=20, pady=8)
        edit_btn.grid(row=0, column=1, padx=5)
        
        # Delete button
        delete_btn = tk.Button(button_frame, text="Delete Student", 
                             command=lambda: self.delete_student_from_details(student),
                             bg="#f44336", fg="white", font=("Arial", 11, "bold"), padx=20, pady=8)
        delete_btn.grid(row=0, column=2, padx=5)
    
    def edit_student(self, tree, students):
        """Edit selected student information"""
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a student to edit")
            return
        
        student_name = selected_item[0]
        student = next((s for s in students if s['full_name'] == student_name), None)
        if not student:
            messagebox.showerror("Error", "Student not found")
            return
        
        # Create edit dialog
        edit_window = tk.Toplevel(self.root)
        edit_window.title("Edit Student Information")
        edit_window.geometry("500x600")
        edit_window.resizable(False, False)
        
        # Center the window
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # Title
        title_label = tk.Label(edit_window, text="Edit Student Information", font=("Arial", 14, "bold"))
        title_label.pack(pady=10)
        
        # Scrollable frame
        canvas = tk.Canvas(edit_window)
        scrollbar = ttk.Scrollbar(edit_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Form fields with current values
        tk.Label(scrollable_frame, text="Full Name:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        name_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        name_entry.insert(0, student.get('full_name', ''))
        name_entry.pack(pady=5, padx=20, fill="x")
        
        # Barangay
        tk.Label(scrollable_frame, text="Barangay:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        barangay_combo = ttk.Combobox(scrollable_frame, values=[
            'A. Bonifacio', 'Abianan', 'Aguinaldo', 'Bagong Sikat', 'Bayanihan', 'Cristo Rey', 'Del Pilar', 'Fatima',
            'Gumamela', 'Holy Child', 'Holy Rosary', 'Immaculate Conception', 'Kalawaan', 'Libertad', 'Little Baguio',
            'Malinis', 'Maluya', 'Mapalad', 'San Antonio', 'San Isidro', 'San Jose', 'San Luis', 'San Miguel',
            'San Rafael', 'San Roque', 'Santa Cruz', 'Santa Lucia', 'Santa Maria', 'Santo Cristo', 'Santo Niño',
            'Santo Rosario', 'Santo Tomas', 'Tibagan', 'Tinigaw', 'Tulay Na Patak', 'Villa Esperanza', 'Villa Jose',
            'Villa Maria'
        ], state="readonly", font=("Arial", 10))
        barangay_combo.set(student.get('barangay', ''))
        barangay_combo.pack(pady=5, padx=20, fill="x")
        
        # Address
        tk.Label(scrollable_frame, text="Address:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        address_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        address_entry.insert(0, student.get('address', ''))
        address_entry.pack(pady=5, padx=20, fill="x")
        
        # Contact Number
        tk.Label(scrollable_frame, text="Contact Number:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        contact_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        contact_entry.insert(0, student.get('contact_number', ''))
        contact_entry.pack(pady=5, padx=20, fill="x")
        
        # School
        tk.Label(scrollable_frame, text="School:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        school_combo = ttk.Combobox(scrollable_frame, values=self.school_options, state="normal", font=("Arial", 10))
        school_combo.set(student.get('school', ''))
        school_combo.pack(pady=5, padx=20, fill="x")
        
        # Course
        tk.Label(scrollable_frame, text="Course:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        course_combo = ttk.Combobox(scrollable_frame, state="normal", font=("Arial", 10))
        course_combo.pack(pady=5, padx=20, fill="x")
        
        # Set initial course options and value based on selected school
        selected_school = student.get('school', '')
        if selected_school in self.course_mapping:
            course_combo['values'] = self.course_mapping[selected_school]
        course_combo.set(student.get('course', ''))
        
        # Bind school selection to update course options
        def update_course_options(event):
            selected_school = school_combo.get()
            if selected_school in self.course_mapping:
                course_combo['values'] = self.course_mapping[selected_school]
                course_combo.set('')  # Clear current selection
            else:
                course_combo['values'] = []
                course_combo.set('')
        
        school_combo.bind('<<ComboboxSelected>>', update_course_options)
        
        # School Year
        tk.Label(scrollable_frame, text="School Year:", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        school_year_entry = tk.Entry(scrollable_frame, font=("Arial", 10))
        school_year_entry.insert(0, student.get('school_year', ''))
        school_year_entry.pack(pady=5, padx=20, fill="x")
        
        # Batch
        tk.Label(scrollable_frame, text="Batch (1-7):", font=("Arial", 10), anchor="w").pack(pady=5, padx=20, fill="x")
        batch_combo = ttk.Combobox(scrollable_frame, values=['1', '2', '3', '4', '5', '6', '7'], state="readonly")
        batch_combo.set(student.get('batch', ''))
        batch_combo.pack(pady=5, padx=20, fill="x")
        
        # Documents section
        tk.Label(scrollable_frame, text="Documents Submitted:", font=("Arial", 12, "bold")).pack(pady=(20, 10), padx=20, anchor="w")
        
        documents = [
            'Certificate of Residency',
            'Pagpapatunay Form',
            'Picture of the House',
            'Good Moral Certificate',
            'Original Certificate of Grades',
            'Proof of Enrollment',
            'School ID'
        ]
        
        doc_vars = {}
        for doc in documents:
            var = tk.BooleanVar(value=student.get('documents', {}).get(doc, False))
            doc_vars[doc] = var
            chk = tk.Checkbutton(scrollable_frame, text=doc, variable=var, font=("Arial", 10), anchor="w")
            chk.pack(anchor="w", padx=40, pady=3, fill="x")
        
        # Button frame
        button_frame = tk.Frame(scrollable_frame)
        button_frame.pack(pady=20, fill="x", padx=20)
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)
        
        def save_changes():
            # Validate required fields
            if not name_entry.get().strip():
                messagebox.showerror("Error", "Full name is required")
                return
            
            school_name = school_combo.get().strip()
            course_name = course_combo.get().strip()
            
            # Add new school to options if entered manually
            if school_name and school_name not in self.school_options:
                self.school_options.append(school_name)
                self.course_mapping[school_name] = []
            
            # Add new course to mapping if entered manually
            if course_name and school_name:
                if course_name not in self.course_mapping.get(school_name, []):
                    self.course_mapping[school_name].append(course_name)
            
            # Save updated options
            self.save_options()
            
            # Update student data
            student['full_name'] = name_entry.get().strip()
            student['barangay'] = barangay_combo.get()
            student['address'] = address_entry.get().strip()
            student['contact_number'] = contact_entry.get().strip()
            student['school'] = school_name
            student['course'] = course_name
            student['school_year'] = school_year_entry.get().strip()
            student['batch'] = batch_combo.get()
            
            # Update documents
            student['documents'] = {doc: var.get() for doc, var in doc_vars.items()}
            
            # Recalculate status
            submitted_docs = sum(var.get() for var in doc_vars.values())
            student['status'] = 'Completed' if submitted_docs == len(documents) else f'Incomplete ({submitted_docs}/{len(documents)})'
            
            # Save to file
            self.save_all_students()
            
            messagebox.showinfo("Success", "Student information updated successfully!")
            edit_window.destroy()
            
            # Refresh the list
            self.show_list_students()
        
        def cancel_edit():
            edit_window.destroy()
        
        # Save and Cancel buttons
        save_btn = tk.Button(button_frame, text="Save Changes", command=save_changes, 
                           bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), padx=20)
        save_btn.grid(row=0, column=0, padx=5, pady=5)
        
        cancel_btn = tk.Button(button_frame, text="Cancel", command=cancel_edit, 
                             bg="#f44336", fg="white", font=("Arial", 10, "bold"), padx=20)
        cancel_btn.grid(row=0, column=1, padx=5, pady=5)
    
    def delete_student(self, tree, students):
        """Move selected student to trash"""
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a student to delete")
            return

        student_name = selected_item[0]

        # Confirm deletion
        if not messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete student '{student_name}'?"):
            return

        # Find and move student to trash
        student_to_remove = next((s for s in self.all_students if s['full_name'] == student_name), None)
        if student_to_remove:
            self.all_students.remove(student_to_remove)
            self.deleted_students.append(student_to_remove)
            self.save_all_students()
            self.save_deleted_students()
            messagebox.showinfo("Success", f"Student '{student_name}' has been moved to Trash")
            self.show_list_students()
        else:
            messagebox.showerror("Error", "Student not found")
    
    def delete_student_from_details(self, student):
        """Move a specific student from details view to trash"""
        student_name = student.get('full_name', 'Unknown')

        # Confirm deletion
        if not messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete student '{student_name}'?"):
            return

        if student in self.all_students:
            self.all_students.remove(student)
            self.deleted_students.append(student)
            self.save_all_students()
            self.save_deleted_students()
            messagebox.showinfo("Success", f"Student '{student_name}' has been moved to Trash")
            self.show_list_students()
        else:
            messagebox.showerror("Error", "Student not found")
    
    def export_to_excel(self, filter_type='all', file_path=None, show_success=True, batch_filter='All'):
        """Export student data to Excel with separate sheets per batch"""
        if filter_type == 'renewed':
            students_to_export = [s for s in self.all_students if 'renewal_date' in s]
        elif filter_type == 'unrenewed':
            students_to_export = [s for s in self.all_students if 'renewal_date' not in s]
        else:
            students_to_export = list(self.all_students)

        if batch_filter != 'All':
            students_to_export = [s for s in students_to_export if s.get('batch', '') == batch_filter]

        if not students_to_export:
            messagebox.showwarning("No Data", "No student data to export for selected filter")
            return None
        
        if file_path is None:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return None
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Group students by batch
            batches = {}
            for student in students_to_export:
                batch = student.get('batch', 'Unknown')
                if batch not in batches:
                    batches[batch] = []
                batches[batch].append(student)

            # Sort students within each batch by full name
            for batch_students in batches.values():
                batch_students.sort(key=lambda s: s.get('full_name', '').strip().lower())

            # Sort all students for 'All Students' tab in selected subset: batches 1-7 first, then others, with names alphabetically within each batch group
            def all_students_sort_key(s):
                batch_raw = s.get('batch', '').strip()
                try:
                    batch_num = int(batch_raw)
                except (TypeError, ValueError):
                    batch_num = None

                if batch_num is None or batch_num < 1 or batch_num > 7:
                    group = 99
                else:
                    group = batch_num

                name = s.get('full_name', '').strip().lower()
                return (group, name)

            sorted_all_students = sorted(students_to_export, key=all_students_sort_key)
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Create sheet for each batch
            for batch in sorted(batches.keys(), key=lambda x: (x.isdigit(), x)):
                ws = wb.create_sheet(f"Batch {batch}")
                
                # Add headers
                headers = ['Full Name', 'Barangay', 'Address', 'Contact Number', 'School', 'Course', 'School Year',
                          'Batch', 'Batch Difference', 'Status', 'Renewed', 'Registration Date', 'Renewal Date']
                
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center_align
                
                # Add data rows
                prev_batch_num = None
                for row_idx, student in enumerate(batches[batch], start=2):
                    renewed_status = "Yes" if 'renewal_date' in student else "No"

                    batch_raw = student.get('batch', '')
                    try:
                        batch_num = int(batch_raw)
                    except (TypeError, ValueError):
                        batch_num = None

                    if prev_batch_num is None or batch_num is None:
                        batch_diff = 0
                    else:
                        batch_diff = batch_num - prev_batch_num

                    prev_batch_num = batch_num
                    
                    ws.cell(row=row_idx, column=1).value = student['full_name']
                    ws.cell(row=row_idx, column=2).value = student.get('barangay', '')
                    ws.cell(row=row_idx, column=3).value = student['address']
                    ws.cell(row=row_idx, column=4).value = student['contact_number']
                    ws.cell(row=row_idx, column=5).value = student['school']
                    ws.cell(row=row_idx, column=6).value = student['course']
                    ws.cell(row=row_idx, column=7).value = student.get('school_year', '')
                    ws.cell(row=row_idx, column=8).value = batch_raw
                    ws.cell(row=row_idx, column=9).value = batch_diff
                    ws.cell(row=row_idx, column=10).value = student.get('status', '')
                    ws.cell(row=row_idx, column=11).value = renewed_status
                    ws.cell(row=row_idx, column=12).value = student.get('registration_date', '')
                    ws.cell(row=row_idx, column=13).value = student.get('renewal_date', '')
                    
                    # Apply borders and alignment to all cells
                    for col in range(1, 14):
                        cell = ws.cell(row=row_idx, column=col)
                        cell.border = border
                        cell.alignment = left_align
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 15  # Barangay
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 20
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 12
                ws.column_dimensions['I'].width = 12
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 12
                ws.column_dimensions['L'].width = 15
                ws.column_dimensions['M'].width = 15
            
            # Create "All Students" sheet
            ws_all = wb.create_sheet("All Students")
            
            # Add headers to All Students sheet (without batch difference)
            all_headers = ['Full Name', 'Barangay', 'Address', 'Contact Number', 'School', 'Course', 'School Year',
                           'Batch', 'Status', 'Renewed', 'Registration Date', 'Renewal Date']
            for col, header in enumerate(all_headers, start=1):
                cell = ws_all.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            # Add all students to All Students sheet (alphabetical order)
            for row_idx, student in enumerate(sorted_all_students, start=2):
                renewed_status = "Yes" if 'renewal_date' in student else "No"
                batch_raw = student.get('batch', '')

                ws_all.cell(row=row_idx, column=1).value = student['full_name']
                ws_all.cell(row=row_idx, column=2).value = student.get('barangay', '')
                ws_all.cell(row=row_idx, column=3).value = student['address']
                ws_all.cell(row=row_idx, column=4).value = student['contact_number']
                ws_all.cell(row=row_idx, column=5).value = student['school']
                ws_all.cell(row=row_idx, column=6).value = student['course']
                ws_all.cell(row=row_idx, column=7).value = student.get('school_year', '')
                ws_all.cell(row=row_idx, column=8).value = batch_raw
                ws_all.cell(row=row_idx, column=9).value = student.get('status', '')
                ws_all.cell(row=row_idx, column=10).value = renewed_status
                ws_all.cell(row=row_idx, column=11).value = student.get('registration_date', '')
                ws_all.cell(row=row_idx, column=12).value = student.get('renewal_date', '')
                
                # Apply borders and alignment to all cells
                for col in range(1, 13):
                    cell = ws_all.cell(row=row_idx, column=col)
                    cell.border = border
                    cell.alignment = left_align
            
            # Adjust column widths for All Students sheet
            ws_all.column_dimensions['A'].width = 20
            ws_all.column_dimensions['B'].width = 15  # Barangay
            ws_all.column_dimensions['C'].width = 25
            ws_all.column_dimensions['D'].width = 15
            ws_all.column_dimensions['E'].width = 20
            ws_all.column_dimensions['F'].width = 20
            ws_all.column_dimensions['G'].width = 15
            ws_all.column_dimensions['H'].width = 12
            ws_all.column_dimensions['I'].width = 12
            ws_all.column_dimensions['J'].width = 15
            ws_all.column_dimensions['K'].width = 12
            ws_all.column_dimensions['L'].width = 15
            
            wb.save(file_path)
            if show_success:
                messagebox.showinfo("Success", f"Excel file exported successfully to {file_path}")
            return file_path
        
        except Exception as e:
            if file_path is not None and not show_success:
                raise
            messagebox.showerror("Error", f"Failed to export: {str(e)}")
            return None

    def print_to_payroll(self):
        """Export student records into PAYROLL_TEMPLATE.xlsx."""
        if not self.all_students:
            messagebox.showwarning("No Data", "No student data to export")
            return

        filter_type = getattr(self, 'payroll_filter_var', tk.StringVar(value='all')).get()
        if filter_type == 'renewed':
            selected = [s for s in self.all_students if 'renewal_date' in s]
        elif filter_type == 'unrenewed':
            selected = [s for s in self.all_students if 'renewal_date' not in s]
        else:
            selected = list(self.all_students)
            selected.sort(key=lambda s: self.get_last_name_sort_key(s.get('full_name', '')))

        batch_filter = getattr(self, 'payroll_batch_var', tk.StringVar(value='All')).get()
        if batch_filter != 'All':
            selected = [s for s in selected if s.get('batch', '') == batch_filter]

        if not selected:
            messagebox.showwarning("No Data", "No student data to export for selected payroll filters")
            return

        template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "PAYROLL_TEMPLATE.xlsx")
        if not os.path.exists(template_path):
            messagebox.showerror("Error", "PAYROLL_TEMPLATE.xlsx was not found in the project folder")
            return

        word_template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "PAYROLL_WORD_TEMPLATE.doc")
        if not os.path.exists(word_template_path):
            messagebox.showerror("Error", "PAYROLL_WORD_TEMPLATE.doc was not found in the project folder")
            return

        selected_path = filedialog.askdirectory(title="Select folder for export files")
        if not selected_path:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        batch_label = "all_batches" if batch_filter == "All" else f"batch_{batch_filter}"
        export_folder = os.path.join(selected_path, f"student_exports_{filter_type}_{batch_label}_{timestamp}")
        os.makedirs(export_folder, exist_ok=True)

        students_path = os.path.join(export_folder, f"student_list_{filter_type}_{batch_label}.xlsx")
        payroll_path = os.path.join(export_folder, f"payroll_{filter_type}_{batch_label}.xlsx")
        word_path = os.path.join(export_folder, f"payroll_{filter_type}_{batch_label}.doc")

        loading_window = self.show_export_loading()
        result_queue = queue.Queue()

        worker = threading.Thread(
            target=self.export_payroll_bundle,
            args=(
                filter_type,
                batch_filter,
                selected,
                template_path,
                word_template_path,
                export_folder,
                students_path,
                payroll_path,
                word_path,
                result_queue
            ),
            daemon=True
        )
        worker.start()
        self.poll_export_result(loading_window, result_queue)

    def export_payroll_bundle(self, filter_type, batch_filter, selected, template_path, word_template_path,
                              export_folder, students_path, payroll_path, word_path, result_queue):
        """Create the student list, payroll Excel, and payroll Word files."""
        try:
            if not self.export_to_excel(filter_type, students_path, show_success=False, batch_filter=batch_filter):
                raise RuntimeError("Student list export was cancelled or failed")

            wb = openpyxl.load_workbook(template_path)
            template_ws = wb.active

            rows_per_page = 15
            start_row = 10
            end_row = 24
            default_pay = 5000
            pages = [selected[i:i + rows_per_page] for i in range(0, len(selected), rows_per_page)]

            page_sheets = [template_ws]
            for _ in range(2, len(pages) + 1):
                page_sheets.append(wb.copy_worksheet(template_ws))

            total_pages = len(pages)
            for page_idx, ws in enumerate(page_sheets, start=1):
                ws.title = f"Payroll - Page {page_idx}"
                ws["O3"] = f"Sheet {page_idx} of {total_pages} Sheets"

            for ws, page_students in zip(page_sheets, pages):
                for row in range(start_row, end_row + 1):
                    ws.cell(row=row, column=2).value = None
                    ws.cell(row=row, column=5).value = None
                    ws.cell(row=row, column=10).value = None

                for row_offset, student in enumerate(page_students):
                    row = start_row + row_offset
                    ws.cell(row=row, column=2).value = student.get("full_name", "")
                    ws.cell(row=row, column=5).value = default_pay
                    ws.cell(row=row, column=10).value = default_pay

                ws.cell(row=start_row + len(page_students), column=2).value = "X-X-X-X"

            wb.save(payroll_path)
            self.export_payroll_word(selected, word_template_path, word_path)
            result_queue.put(("success", export_folder, students_path, payroll_path, word_path))

        except Exception as e:
            result_queue.put(("error", str(e)))

    def show_export_loading(self):
        """Show a small loading dialog while export files are being created."""
        loading_window = tk.Toplevel(self.root)
        loading_window.title("Exporting")
        loading_window.geometry("320x120")
        loading_window.resizable(False, False)
        loading_window.transient(self.root)
        loading_window.grab_set()
        loading_window.protocol("WM_DELETE_WINDOW", lambda: None)

        tk.Label(
            loading_window,
            text="Creating export files...",
            font=("Arial", 11, "bold")
        ).pack(pady=(18, 8))

        progress = ttk.Progressbar(loading_window, mode="indeterminate", length=240)
        progress.pack(pady=8)
        progress.start(12)

        tk.Label(
            loading_window,
            text="Please wait while Excel and Word files are generated.",
            font=("Arial", 9)
        ).pack(pady=(4, 0))

        loading_window.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 160
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 60
        loading_window.geometry(f"+{x}+{y}")
        return loading_window

    def poll_export_result(self, loading_window, result_queue):
        """Poll a worker-thread result and report it safely on the Tk thread."""
        try:
            result = result_queue.get_nowait()
        except queue.Empty:
            self.root.after(100, lambda: self.poll_export_result(loading_window, result_queue))
            return

        if loading_window.winfo_exists():
            loading_window.grab_release()
            loading_window.destroy()

        if result[0] == "success":
            _, export_folder, students_path, payroll_path, word_path = result
            messagebox.showinfo(
                "Success",
                f"Export folder created successfully:\n{export_folder}\n\n"
                f"Files created:\n{os.path.basename(students_path)}\n{os.path.basename(payroll_path)}"
                f"\n{os.path.basename(word_path)}"
            )
        else:
            messagebox.showerror("Error", f"Failed to create payroll export: {result[1]}")

    def export_payroll_word(self, selected, template_path, output_path):
        """Export student records into the payroll Word table template."""
        try:
            import pythoncom
            import win32com.client
        except ImportError as exc:
            raise RuntimeError("pywin32 is required to export the Word payroll file") from exc

        word = None
        doc = None
        try:
            pythoncom.CoInitialize()
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(template_path, False, True)

            rows_per_page = 15
            word_students = sorted(selected, key=self.get_word_payroll_sort_key)
            pages = [word_students[i:i + rows_per_page] for i in range(0, len(word_students), rows_per_page)]
            if not pages:
                pages = [[]]

            template_range = doc.Content
            template_range.Copy()

            for _ in range(2, len(pages) + 1):
                end_range = doc.Range(doc.Content.End - 1, doc.Content.End - 1)
                end_range.InsertBreak(7)  # Word page break
                paste_range = doc.Range(doc.Content.End - 1, doc.Content.End - 1)
                paste_range.Paste()

            for page_idx, page_students in enumerate(pages, start=1):
                table = doc.Tables(page_idx)

                for row in range(2, 17):
                    self.set_word_table_cell(table, row, 2, "")
                    self.set_word_table_cell(table, row, 3, "")
                    self.set_word_table_cell(table, row, 4, "")

                self.set_word_table_cell(table, 17, 2, "")

                for row_offset, student in enumerate(page_students):
                    row = 2 + row_offset
                    self.set_word_table_cell(table, row, 2, student.get("full_name", "").upper())
                    self.set_word_table_cell(table, row, 3, self.format_year_level(student.get("school_year", "")))
                    self.set_word_table_cell(table, row, 4, self.format_word_school_name(student.get("school", "")))

                self.set_word_table_cell(table, 2 + len(page_students), 2, "X-X-X-X")

            doc.SaveAs2(output_path, FileFormat=0)

        finally:
            if doc is not None:
                doc.Close(False)
            if word is not None:
                word.Quit()
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def set_word_table_cell(self, table, row, column, value):
        """Set a Word table cell while tolerating merged template cells."""
        try:
            table.Cell(row, column).Range.Text = value
            return True
        except Exception:
            return False

    def format_year_level(self, school_year):
        """Format the year level for the Word payroll table."""
        year = str(school_year).strip()
        if not year:
            return ""

        digits = "".join(ch for ch in year if ch.isdigit())
        if digits:
            year = digits

        ordinal_suffixes = {
            "1": "ST",
            "2": "ND",
            "3": "RD"
        }
        if year in ordinal_suffixes or year.isdigit():
            suffix = ordinal_suffixes.get(year, "TH")
            return f"{year}{suffix} YEAR"

        if year.upper().endswith("YEAR"):
            return year.upper()
        return f"{year.upper()} YEAR"

    def format_word_school_name(self, school_name):
        """Format the school name for the Word payroll table."""
        shortened = re.sub(r"\bSan\s+Jose\s+City\b", "", school_name, flags=re.IGNORECASE)
        shortened = " ".join(shortened.split())
        return shortened.upper()

    def get_word_payroll_sort_key(self, student):
        """Sort Word payroll rows by year level, then student last name."""
        year_raw = str(student.get("school_year", "")).strip()
        digits = "".join(ch for ch in year_raw if ch.isdigit())
        try:
            year_num = int(digits)
        except ValueError:
            year_num = 99

        return (year_num, self.get_last_name_sort_key(student.get("full_name", "")))

    def get_last_name_sort_key(self, full_name):
        """Return a sort key that prioritizes the student's last name."""
        name = full_name.strip().lower()
        if not name:
            return ("", "")

        if "," in name:
            last_name, rest = name.split(",", 1)
            return (last_name.strip(), rest.strip())

        parts = name.split()
        if len(parts) == 1:
            return (parts[0], "")

        return (parts[-1], " ".join(parts[:-1]))

# Create main window
root = tk.Tk()
root.title("Student Information System")
root.geometry("600x700")

# Initialize application
app = StudentInfoSystem(root)

# Run the application
root.mainloop()
